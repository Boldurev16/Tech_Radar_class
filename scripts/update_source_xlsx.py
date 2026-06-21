# -*- coding: utf-8 -*-
"""Write classifier predictions back to data/source.xlsx (target sheet only)."""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from classifier import TechRadarClassifier
from rules import NN_SPUTNIK_PATTERNS

TARGET_SHEET = "Build your Technology Radar"
SOURCE_PATH = ROOT / "data" / "source.xlsx"
BACKUP_PATH = ROOT / "data" / "source_backup.xlsx"
LOW_CONF_PATH = ROOT / "output" / "low_confidence_records.csv"

COL_QUADRANT = "quadrant"
COL_BLOCK = "block"
COL_NAME = "name"
COL_DESCRIPTION = "description"
COL_RING = "ring"


def snapshot_other_sheets(wb, target_sheet: str) -> dict[str, list[list[object]]]:
    snapshot: dict[str, list[list[object]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == target_sheet:
            continue
        ws = wb[sheet_name]
        snapshot[sheet_name] = [[cell.value for cell in row] for row in ws.iter_rows()]
    return snapshot


def verify_other_sheets_intact(
    snapshot_before: dict[str, list[list[object]]],
    wb_after,
    target_sheet: str,
) -> None:
    for sheet_name, vals_before in snapshot_before.items():
        if sheet_name not in wb_after.sheetnames:
            raise AssertionError(f"ОШИБКА: лист '{sheet_name}' пропал из файла!")
        ws_after = wb_after[sheet_name]
        vals_after = [[cell.value for cell in row] for row in ws_after.iter_rows()]
        assert vals_before == vals_after, f"ОШИБКА: лист '{sheet_name}' был изменён!"
    print("[verify] Все остальные листы нетронуты ✓")


def update_source_xlsx() -> None:
    if not SOURCE_PATH.exists():
        raise FileNotFoundError(f"Файл не найден: {SOURCE_PATH}")

    BACKUP_PATH.parent.mkdir(parents=True, exist_ok=True)
    LOW_CONF_PATH.parent.mkdir(parents=True, exist_ok=True)

    shutil.copy2(SOURCE_PATH, BACKUP_PATH)
    print(f"[backup] Сохранена резервная копия: {BACKUP_PATH}")

    df = pd.read_excel(SOURCE_PATH, sheet_name=TARGET_SHEET)
    print(f"[read] Лист '{TARGET_SHEET}': {len(df)} строк, колонки: {list(df.columns)}")

    required = [COL_NAME, COL_DESCRIPTION, COL_QUADRANT, COL_BLOCK]
    missing = [col for col in required if col not in df.columns]
    if missing:
        raise ValueError(f"Колонки не найдены: {missing}. Доступные: {list(df.columns)}")

    wb = load_workbook(SOURCE_PATH)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(
            f"Лист '{TARGET_SHEET}' не найден. Листы в файле: {wb.sheetnames}"
        )
    other_sheets_snapshot = snapshot_other_sheets(wb, TARGET_SHEET)

    clf = TechRadarClassifier(rebuild_prototypes=False)
    results: list[dict] = []

    print(f"[classify] Классификация {len(df)} строк...")
    for _, row in df.iterrows():
        name = str(row[COL_NAME]) if pd.notna(row[COL_NAME]) else ""
        desc = str(row[COL_DESCRIPTION]) if pd.notna(row[COL_DESCRIPTION]) else ""
        ring = str(row[COL_RING]) if COL_RING in df.columns and pd.notna(row.get(COL_RING)) else ""
        raw_block = str(row[COL_BLOCK]) if pd.notna(row[COL_BLOCK]) else ""

        if not name.strip() or name.strip().lower() in ("nan", "name", "key"):
            results.append(
                {
                    "quadrant_new": row[COL_QUADRANT],
                    "block_new": row[COL_BLOCK],
                    "skipped": True,
                }
            )
            continue

        result = clf.classify(name, desc, ring=ring, raw_block=raw_block)

        is_sputnik = any(pattern in raw_block for pattern in NN_SPUTNIK_PATTERNS)
        if is_sputnik and result.block_confidence < 0.5:
            block_to_write = raw_block
        else:
            block_to_write = result.block

        results.append(
            {
                "quadrant_new": result.quadrant,
                "block_new": block_to_write,
                "quadrant_conf": result.quadrant_confidence,
                "block_conf": result.block_confidence,
                "skipped": False,
            }
        )

    ws = wb[TARGET_SHEET]
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    col_index = {cell: idx + 1 for idx, cell in enumerate(header_row) if cell}

    if COL_QUADRANT not in col_index:
        raise ValueError(
            f"Колонка '{COL_QUADRANT}' не найдена в заголовке листа. Заголовок: {header_row}"
        )
    if COL_BLOCK not in col_index:
        raise ValueError(
            f"Колонка '{COL_BLOCK}' не найдена в заголовке листа. Заголовок: {header_row}"
        )

    q_col = col_index[COL_QUADRANT]
    b_col = col_index[COL_BLOCK]

    updated_count = 0
    skipped_count = 0

    for i, res in enumerate(results):
        excel_row = i + 2
        if res["skipped"]:
            skipped_count += 1
            continue
        ws.cell(row=excel_row, column=q_col).value = res["quadrant_new"]
        ws.cell(row=excel_row, column=b_col).value = res["block_new"]
        updated_count += 1

    wb.save(SOURCE_PATH)
    print(f"[done] Обновлено {updated_count} строк, пропущено {skipped_count}")
    print(f"[done] Файл сохранён: {SOURCE_PATH}")
    print(f"[done] Резервная копия: {BACKUP_PATH}")

    wb_after = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    verify_other_sheets_intact(other_sheets_snapshot, wb_after, TARGET_SHEET)
    wb_after.close()

    low_conf = [
        (df.iloc[i][COL_NAME], r["quadrant_conf"], r["block_conf"])
        for i, r in enumerate(results)
        if not r["skipped"]
        and (r.get("quadrant_conf", 1) < 0.6 or r.get("block_conf", 1) < 0.6)
    ]
    if low_conf:
        report_df = pd.DataFrame(
            low_conf, columns=["name", "quadrant_conf", "block_conf"]
        ).sort_values("quadrant_conf")
        report_df.to_csv(LOW_CONF_PATH, index=False, encoding="utf-8-sig")
        print(
            f"[report] {len(low_conf)} записей с conf < 0.6 → {LOW_CONF_PATH}"
        )
    else:
        print("[report] Записей с conf < 0.6 не найдено")


if __name__ == "__main__":
    update_source_xlsx()
