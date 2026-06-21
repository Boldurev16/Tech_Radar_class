# -*- coding: utf-8 -*-
"""Compare manual markup with classifier output and merge manual into source.xlsx."""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

MANUAL_PATH = ROOT / "data" / "source_16.06.xlsx"
AUTO_PATH = ROOT / "data" / "source.xlsx"
BACKUP_PATH = ROOT / "data" / "source_backup_before_merge.xlsx"
DIFF_REPORT_PATH = ROOT / "output" / "diff_report.xlsx"
TARGET_SHEET = "Build your Technology Radar"

KEY_COL = "name"
COL_QUADRANT = "quadrant"
COL_BLOCK = "block"

FILL_BOTH = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
FILL_QUADRANT = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
FILL_BLOCK = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")

STATUS_FILLS = {
    "BOTH_DIFF": FILL_BOTH,
    "QUADRANT_DIFF": FILL_QUADRANT,
    "BLOCK_DIFF": FILL_BLOCK,
}


def normalize_value(value: object) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    text = str(value).strip()
    return "" if text.lower() == "nan" else text


def classify_diff(row: pd.Series) -> str:
    q_match = normalize_value(row["quadrant_manual"]) == normalize_value(row["quadrant_auto"])
    b_match = normalize_value(row["block_manual"]) == normalize_value(row["block_auto"])
    if q_match and b_match:
        return "MATCH"
    if q_match and not b_match:
        return "BLOCK_DIFF"
    if not q_match and b_match:
        return "QUADRANT_DIFF"
    return "BOTH_DIFF"


def dedupe_by_key(df: pd.DataFrame) -> pd.DataFrame:
    return df.drop_duplicates(subset=[KEY_COL], keep="last").reset_index(drop=True)


def snapshot_other_sheets(wb, target_sheet: str) -> dict[str, list[list[object]]]:
    snapshot: dict[str, list[list[object]]] = {}
    for sheet_name in wb.sheetnames:
        if sheet_name == target_sheet:
            continue
        ws = wb[sheet_name]
        snapshot[sheet_name] = [[cell.value for cell in row] for row in ws.iter_rows()]
    return snapshot


def verify_other_sheets_intact(
    snapshot_before: dict[str, list[list[object]]],
    wb_after,
    target_sheet: str,
) -> None:
    for sheet_name, vals_before in snapshot_before.items():
        if sheet_name not in wb_after.sheetnames:
            raise AssertionError(f"ОШИБКА: лист '{sheet_name}' пропал из файла!")
        ws_after = wb_after[sheet_name]
        vals_after = [[cell.value for cell in row] for row in ws_after.iter_rows()]
        assert vals_before == vals_after, f"ОШИБКА: лист '{sheet_name}' был изменён!"
    print("[verify] Все остальные листы source.xlsx нетронуты ✓")


def build_comparison(df_manual: pd.DataFrame, df_auto: pd.DataFrame) -> pd.DataFrame:
    manual = dedupe_by_key(df_manual[[KEY_COL, COL_QUADRANT, COL_BLOCK]])
    auto = dedupe_by_key(df_auto[[KEY_COL, COL_QUADRANT, COL_BLOCK]])

    df = manual.merge(
        auto,
        on=KEY_COL,
        suffixes=("_manual", "_auto"),
        how="outer",
        indicator=True,
    )
    df["diff_status"] = df.apply(classify_diff, axis=1)

    only_manual = df["_merge"] == "left_only"
    only_auto = df["_merge"] == "right_only"
    if only_manual.any():
        df.loc[only_manual, "diff_status"] = "ONLY_MANUAL"
    if only_auto.any():
        df.loc[only_auto, "diff_status"] = "ONLY_AUTO"

    return df.drop(columns=["_merge"])


def save_diff_report(df: pd.DataFrame) -> None:
    DIFF_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)

    total = len(df)
    status_counts = df["diff_status"].value_counts()
    summary_rows = []
    for status in ["MATCH", "QUADRANT_DIFF", "BLOCK_DIFF", "BOTH_DIFF", "ONLY_MANUAL", "ONLY_AUTO"]:
        count = int(status_counts.get(status, 0))
        pct = round(count / total * 100, 2) if total else 0.0
        summary_rows.append({"Статус": status, "Количество": count, "% от total": pct})

    summary_df = pd.DataFrame(summary_rows)
    diffs_df = df[df["diff_status"] != "MATCH"].copy()
    diffs_df = diffs_df[
        [KEY_COL, "quadrant_manual", "quadrant_auto", "block_manual", "block_auto", "diff_status"]
    ]

    with pd.ExcelWriter(DIFF_REPORT_PATH, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Сводка", index=False)
        diffs_df.to_excel(writer, sheet_name="Расхождения", index=False)

        ws = writer.sheets["Расхождения"]
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=6).value
            fill = STATUS_FILLS.get(str(status))
            if fill is None:
                continue
            for col_idx in range(1, 7):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    print(f"[report] Отчёт сохранён: {DIFF_REPORT_PATH}")


def update_source_from_manual(df_manual: pd.DataFrame) -> tuple[int, list[str]]:
    if not AUTO_PATH.exists():
        raise FileNotFoundError(f"Файл не найден: {AUTO_PATH}")

    shutil.copy2(AUTO_PATH, BACKUP_PATH)
    print(f"[backup] Резервная копия: {BACKUP_PATH}")

    wb = load_workbook(AUTO_PATH)
    if TARGET_SHEET not in wb.sheetnames:
        raise ValueError(f"Лист '{TARGET_SHEET}' не найден. Листы: {wb.sheetnames}")

    other_sheets_snapshot = snapshot_other_sheets(wb, TARGET_SHEET)
    ws = wb[TARGET_SHEET]

    header_row = next(ws.iter_rows(min_row=1, max_row=1))
    header = {cell.value: cell.column for cell in header_row if cell.value}

    for col_name in (KEY_COL, COL_QUADRANT, COL_BLOCK):
        if col_name not in header:
            raise ValueError(f"Колонка '{col_name}' не найдена. Заголовок: {list(header.keys())}")

    key_col = header[KEY_COL]
    q_col = header[COL_QUADRANT]
    b_col = header[COL_BLOCK]

    manual_lookup = {
        normalize_value(row[KEY_COL]): (row[COL_QUADRANT], row[COL_BLOCK])
        for _, row in df_manual.iterrows()
        if normalize_value(row[KEY_COL])
    }

    updated = 0
    not_found: list[str] = []

    for row in ws.iter_rows(min_row=2):
        key_val = normalize_value(row[key_col - 1].value)
        if not key_val:
            continue
        if key_val in manual_lookup:
            q_new, b_new = manual_lookup[key_val]
            row[q_col - 1].value = q_new
            row[b_col - 1].value = b_new
            updated += 1
        else:
            not_found.append(key_val)

    wb.save(AUTO_PATH)

    wb_after = load_workbook(AUTO_PATH, read_only=True, data_only=True)
    verify_other_sheets_intact(other_sheets_snapshot, wb_after, TARGET_SHEET)
    wb_after.close()

    return updated, not_found


def print_final_summary(df: pd.DataFrame, updated: int, not_found: list[str]) -> None:
    total = len(df)
    counts = df["diff_status"].value_counts()
    mismatches = total - int(counts.get("MATCH", 0))

    print("\n" + "=" * 60)
    print("ИТОГО: сравнение ручной разметки и классификатора")
    print("=" * 60)
    print(f"  Эталон:     {MANUAL_PATH.name}")
    print(f"  Авто:       {AUTO_PATH.name}")
    print(f"  Всего строк (outer join по name): {total}")
    print(f"  MATCH:              {int(counts.get('MATCH', 0))}")
    print(f"  QUADRANT_DIFF:      {int(counts.get('QUADRANT_DIFF', 0))}")
    print(f"  BLOCK_DIFF:         {int(counts.get('BLOCK_DIFF', 0))}")
    print(f"  BOTH_DIFF:          {int(counts.get('BOTH_DIFF', 0))}")
    print(f"  ONLY_MANUAL:        {int(counts.get('ONLY_MANUAL', 0))}")
    print(f"  ONLY_AUTO:          {int(counts.get('ONLY_AUTO', 0))}")
    print(f"  Расхождений (≠ MATCH): {mismatches} ({round(mismatches / total * 100, 1)}%)")
    print("-" * 60)
    print(f"  Обновлено в source.xlsx: {updated} строк (приоритет — ручная разметка)")
    print(f"  Отчёт: {DIFF_REPORT_PATH}")
    print(f"  Backup: {BACKUP_PATH}")
    if not_found:
        preview = ", ".join(not_found[:10])
        suffix = "..." if len(not_found) > 10 else ""
        print(f"  [warn] Не найдены в эталоне ({len(not_found)}): {preview}{suffix}")
    print("=" * 60)


def main() -> None:
    df_manual = pd.read_excel(MANUAL_PATH, sheet_name=TARGET_SHEET)
    df_auto = pd.read_excel(AUTO_PATH, sheet_name=TARGET_SHEET)

    print(f"[read] manual: {len(df_manual)} строк, auto: {len(df_auto)} строк")
    print(f"[read] колонки: {list(df_manual.columns)}")

    df_compare = build_comparison(df_manual, df_auto)
    save_diff_report(df_compare)

    updated, not_found = update_source_from_manual(df_manual)
    print(f"[done] Обновлено: {updated} строк в {AUTO_PATH}")

    print_final_summary(df_compare, updated, not_found)


if __name__ == "__main__":
    main()
