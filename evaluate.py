# -*- coding: utf-8 -*-
"""Evaluation metrics, diagnostics, and export for Tech Radar classifier."""
from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import TYPE_CHECKING, Any

import numpy as np
import pandas as pd
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics import accuracy_score, confusion_matrix, f1_score
from sklearn.model_selection import StratifiedShuffleSplit

from rules import (
    BLOCK_LABELS,
    BLOCK_ALIASES,
    QUADRANT_LABELS,
    canonical_block,
    canonical_quadrant,
    is_nn_sputnik_block,
    normalize_block,
)
from semantic import build_training_corpus

if TYPE_CHECKING:
    from classifier import TechRadarClassifier

DATA_PATH = Path(__file__).resolve().parent / "data" / "source.xlsx"
MODELS_DIR = Path(__file__).resolve().parent / "models"
HOLDOUT_RATIO = 0.2
RANDOM_SEED = 42
MIN_STRATUM_SIZE = 5
ENSEMBLE_WEIGHTS_PATH = MODELS_DIR / "ensemble_weights.json"
DIAGNOSTICS_PATH = Path(__file__).resolve().parent / "diagnostics_report.json"
UNMAPPED_BLOCKS_PATH = Path(__file__).resolve().parent / "unmapped_blocks.csv"
RESULTS_DIR = Path(__file__).resolve().parent / "results"
ITERATION3_METRICS_PATH = RESULTS_DIR / "iteration3_metrics.json"


def exclude_nn_sputnik_rows(df: pd.DataFrame, block_col: str = "block") -> pd.DataFrame:
    if block_col not in df.columns:
        return df.copy()
    mask_sputnik = df[block_col].apply(is_nn_sputnik_block)
    n_removed = int(mask_sputnik.sum())
    if n_removed:
        print(f"[load_dataset] Исключено {n_removed} записей НН-Спутник из {len(df)} total")
    return df.loc[~mask_sputnik].reset_index(drop=True)


def load_dataset(path: str | Path | None = None, raw_blocks: bool = False) -> pd.DataFrame:
    source = Path(path or DATA_PATH)
    raw = pd.read_excel(source, sheet_name="Build your Technology Radar")
    raw = exclude_nn_sputnik_rows(raw, block_col="block")
    if raw_blocks:
        corpus = raw.rename(columns={"quadrant": "quadrant_raw", "block": "block_raw"})
        corpus["quadrant"] = corpus["quadrant_raw"].map(canonical_quadrant)
        corpus["block"] = corpus["block_raw"].map(
            lambda x: normalize_block(x) if pd.notna(x) else None
        )
        corpus = corpus.dropna(subset=["name", "quadrant", "block"])
        corpus["text"] = corpus.apply(
            lambda r: f"{str(r['name']).lower()} | {str(r.get('description', '') or '').lower()}",
            axis=1,
        )
        return corpus.reset_index(drop=True)
    return build_training_corpus(raw)


def load_full_export_dataset(path: str | Path | None = None) -> pd.DataFrame:
    """Full source rows for batch export (includes НН-Спутnik metatag rows)."""
    source = Path(path or DATA_PATH)
    raw = pd.read_excel(source, sheet_name="Build your Technology Radar")
    return raw[["name", "description", "quadrant", "block", "ring"]].copy()


def load_block_process_codes(path: str | Path | None = None) -> dict[str, list[str]]:
    source = Path(path or DATA_PATH)
    block_sheet = pd.read_excel(source, sheet_name="Block")
    mapping: dict[str, list[str]] = {}
    for _, row in block_sheet.iterrows():
        label = str(row.get("label", "") or "").replace("\u200b", "").strip()
        if not label:
            continue
        processes = str(row.get("Процессы", "") or "")
        clean_codes = [
            match.group(1).lower()
            for match in re.finditer(r"\[([A-Z]{2,5}\.\d+(?:\.\d+)*)\]", processes)
        ]
        if clean_codes:
            mapping[label] = clean_codes
    return mapping


def train_test_split_df(
    df: pd.DataFrame,
    test_ratio: float = HOLDOUT_RATIO,
    seed: int = RANDOM_SEED,
    stratify_col: str = "quadrant",
) -> tuple[pd.DataFrame, pd.DataFrame]:
    if df.empty:
        return df.copy(), df.copy()

    df = df.reset_index(drop=True)
    if stratify_col == "multi":
        labels = df["quadrant"].astype(str) + "|||" + df["block"].astype(str)
    else:
        labels = df[stratify_col].astype(str)
    counts = labels.value_counts()
    rare_mask = labels.map(counts) < MIN_STRATUM_SIZE
    rare_idx = df.index[rare_mask].tolist()
    common_df = df.loc[~rare_mask].reset_index(drop=True)

    if common_df.empty:
        return df.copy(), df.iloc[0:0].copy()

    if stratify_col == "multi":
        common_labels = (
            common_df["quadrant"].astype(str) + "|||" + common_df["block"].astype(str)
        )
    else:
        common_labels = common_df[stratify_col].astype(str)

    splitter = StratifiedShuffleSplit(n_splits=1, test_size=test_ratio, random_state=seed)
    train_idx, test_idx = next(splitter.split(common_df, common_labels))
    train_df = common_df.iloc[train_idx].copy()
    test_df = common_df.iloc[test_idx].copy()

    if rare_idx:
        rare_df = df.loc[rare_idx]
        train_df = pd.concat([train_df, rare_df], ignore_index=True)

    return train_df.reset_index(drop=True), test_df.reset_index(drop=True)


def extract_keywords_for_class(
    df: pd.DataFrame,
    class_name: str,
    top_n: int = 20,
    label_col: str = "quadrant",
) -> list[str]:
    if label_col not in df.columns:
        raise ValueError(f"Column {label_col} not in dataframe")

    texts = df.apply(
        lambda r: f"{r.get('name', '')} {r.get('description', '')}".lower(),
        axis=1,
    ).astype(str)
    y = (df[label_col] == class_name).astype(int)
    if y.sum() < 2:
        return []

    vectorizer = TfidfVectorizer(
        analyzer="word",
        ngram_range=(1, 2),
        min_df=1,
        max_df=0.95,
        token_pattern=r"(?u)\b[\w\-]{3,}\b",
    )
    x = vectorizer.fit_transform(texts)
    pos_mask = y.to_numpy().astype(bool)
    pos_mean = np.asarray(x[pos_mask].mean(axis=0)).ravel()
    neg_mean = np.asarray(x[~pos_mask].mean(axis=0)).ravel()
    delta = pos_mean - neg_mean
    terms = vectorizer.get_feature_names_out()
    ranked = sorted(zip(terms, delta), key=lambda item: -item[1])
    return [term for term, score in ranked[:top_n] if score > 0]


def log_unmapped_blocks(path: str | Path | None = None) -> pd.DataFrame:
    source = Path(path or DATA_PATH)
    raw = pd.read_excel(source, sheet_name="Build your Technology Radar")
    rows: list[dict[str, Any]] = []
    for value, count in raw["block"].value_counts().items():
        raw_value = str(value).strip()
        cleaned = raw_value.replace("\u200b", "")
        if cleaned in BLOCK_LABELS:
            continue
        if cleaned in BLOCK_ALIASES:
            continue
        if is_nn_sputnik_block(raw_value):
            continue
        mapped = normalize_block(raw_value, use_semantic=False)
        if mapped is None:
            rows.append(
                {
                    "raw_block": raw_value,
                    "count": int(count),
                    "semantic_match": mapped or "",
                }
            )
    out = pd.DataFrame(rows)
    if not out.empty:
        out.to_csv(UNMAPPED_BLOCKS_PATH, index=False, encoding="utf-8-sig")
    else:
        pd.DataFrame(columns=["raw_block", "count", "semantic_match"]).to_csv(
            UNMAPPED_BLOCKS_PATH, index=False, encoding="utf-8-sig"
        )
    return out


def class_distribution(df: pd.DataFrame) -> dict[str, dict[str, int]]:
    return {
        "quadrant": dict(Counter(df["quadrant"])),
        "block": dict(Counter(df["block"])),
    }


def compatibility_matrix(
    df: pd.DataFrame,
    row_col: str = "quadrant",
    col_col: str = "block",
) -> pd.DataFrame:
    pivot = pd.crosstab(df[row_col], df[col_col], normalize="index")
    pivot = pivot.reindex(columns=sorted(pivot.columns), fill_value=0.0)
    return pivot


def block_quadrant_matrix(df: pd.DataFrame) -> pd.DataFrame:
    pivot = pd.crosstab(df["block"], df["quadrant"], normalize="index")
    pivot = pivot.reindex(columns=sorted(pivot.columns), fill_value=0.0)
    return pivot


def _per_class_f1(y_true: list[str], y_pred: list[str], labels: list[str]) -> dict[str, float]:
    scores: dict[str, float] = {}
    for label in labels:
        tp = sum(1 for t, p in zip(y_true, y_pred) if t == label and p == label)
        fp = sum(1 for t, p in zip(y_true, y_pred) if t != label and p == label)
        fn = sum(1 for t, p in zip(y_true, y_pred) if t == label and p != label)
        denom = 2 * tp + fp + fn
        scores[label] = float(2 * tp / denom) if denom else 0.0
    return scores


def _top_confusion_pairs(
    y_true: list[str],
    y_pred: list[str],
    top_n: int = 10,
) -> list[dict[str, Any]]:
    pairs: Counter[tuple[str, str]] = Counter()
    for true_label, pred_label in zip(y_true, y_pred):
        if true_label != pred_label:
            pairs[(true_label, pred_label)] += 1
    return [
        {"true": true, "pred": pred, "count": count}
        for (true, pred), count in pairs.most_common(top_n)
    ]


def full_diagnostics(clf: "TechRadarClassifier", df_test: pd.DataFrame) -> dict[str, Any]:
    preds = clf.classify_batch(df_test[["name", "description"]])
    y_q_true = [canonical_quadrant(x) for x in df_test["quadrant"]]
    y_b_true = [canonical_block(x) for x in df_test["block"]]
    y_q_pred = preds["quadrant"].tolist()
    y_b_pred = preds["block"].tolist()

    q_labels = sorted(set(y_q_true) | set(y_q_pred))
    b_labels = sorted(set(y_b_true) | set(y_b_pred))
    q_f1 = _per_class_f1(y_q_true, y_q_pred, q_labels)
    b_f1 = _per_class_f1(y_b_true, y_b_pred, b_labels)

    misclassified: list[dict[str, Any]] = []
    for idx, row in df_test.reset_index(drop=True).iterrows():
        pred = preds.iloc[idx]
        tq = canonical_quadrant(row["quadrant"])
        tb = canonical_block(row["block"])
        if pred["quadrant"] != tq or pred["block"] != tb:
            misclassified.append(
                {
                    "name": row["name"],
                    "quadrant_true": tq,
                    "quadrant_pred": pred["quadrant"],
                    "block_true": tb,
                    "block_pred": pred["block"],
                    "quadrant_top3": pred["quadrant_top3"],
                    "block_top3": pred["block_top3"],
                }
            )

    q_correct_conf = [
        float(p["quadrant_confidence"])
        for p, t in zip(preds.to_dict("records"), y_q_true)
        if p["quadrant"] == t
    ]
    q_wrong_conf = [
        float(p["quadrant_confidence"])
        for p, t in zip(preds.to_dict("records"), y_q_true)
        if p["quadrant"] != t
    ]
    b_correct_conf = [
        float(p["block_confidence"])
        for p, t in zip(preds.to_dict("records"), y_b_true)
        if p["block"] == t
    ]
    b_wrong_conf = [
        float(p["block_confidence"])
        for p, t in zip(preds.to_dict("records"), y_b_true)
        if p["block"] != t
    ]

    report = {
        "samples": len(df_test),
        "quadrant_per_class_f1": dict(sorted(q_f1.items(), key=lambda x: x[1])),
        "block_per_class_f1": dict(sorted(b_f1.items(), key=lambda x: x[1])),
        "quadrant_confusion_pairs_top10": _top_confusion_pairs(y_q_true, y_q_pred, 10),
        "block_confusion_pairs_top10": _top_confusion_pairs(y_b_true, y_b_pred, 10),
        "misclassified_examples": misclassified[:50],
        "confidence_distribution": {
            "quadrant_correct_mean": float(np.mean(q_correct_conf)) if q_correct_conf else 0.0,
            "quadrant_wrong_mean": float(np.mean(q_wrong_conf)) if q_wrong_conf else 0.0,
            "block_correct_mean": float(np.mean(b_correct_conf)) if b_correct_conf else 0.0,
            "block_wrong_mean": float(np.mean(b_wrong_conf)) if b_wrong_conf else 0.0,
        },
        "low_f1_quadrant_classes": [k for k, v in q_f1.items() if v < 0.3],
        "low_f1_block_classes": [k for k, v in b_f1.items() if v < 0.3],
    }
    return report


def save_diagnostics_report(report: dict[str, Any], path: str | Path | None = None) -> Path:
    target = Path(path or DIAGNOSTICS_PATH)
    with target.open("w", encoding="utf-8") as fh:
        json.dump(report, fh, ensure_ascii=False, indent=2)
    return target


def evaluate(df_labeled: pd.DataFrame, classifier: "TechRadarClassifier") -> dict:
    required = {"name", "description", "quadrant", "block"}
    missing = required - set(df_labeled.columns)
    if missing:
        raise ValueError(f"Missing columns: {sorted(missing)}")

    preds = classifier.classify_batch(
        df_labeled[[c for c in ("name", "description", "block", "ring") if c in df_labeled.columns]]
    )
    y_q_true = [canonical_quadrant(x) for x in df_labeled["quadrant"]]
    y_b_true = [canonical_block(x) for x in df_labeled["block"]]
    y_q_pred = preds["quadrant"].tolist()
    y_b_pred = preds["block"].tolist()

    valid_mask = [
        tq is not None and tb is not None for tq, tb in zip(y_q_true, y_b_true)
    ]
    y_q_true = [x for x, ok in zip(y_q_true, valid_mask) if ok]
    y_b_true = [x for x, ok in zip(y_b_true, valid_mask) if ok]
    y_q_pred = [x for x, ok in zip(y_q_pred, valid_mask) if ok]
    y_b_pred = [x for x, ok in zip(y_b_pred, valid_mask) if ok]
    q_conf = preds.loc[valid_mask, "quadrant_confidence"].tolist()
    b_conf = preds.loc[valid_mask, "block_confidence"].tolist()

    q_labels = sorted(set(y_q_true) | set(y_q_pred))
    b_labels = sorted(set(y_b_true) | set(y_b_pred))

    cm_q = pd.DataFrame(
        confusion_matrix(y_q_true, y_q_pred, labels=q_labels),
        index=q_labels,
        columns=q_labels,
    )
    cm_b = pd.DataFrame(
        confusion_matrix(y_b_true, y_b_pred, labels=b_labels),
        index=b_labels,
        columns=b_labels,
    )

    joint = [
        int(pq == tq and pb == tb)
        for pq, pb, tq, tb in zip(y_q_pred, y_b_pred, y_q_true, y_b_true)
    ]
    low_conf = [
        1
        for qc, bc in zip(q_conf, b_conf)
        if qc < 0.6 or bc < 0.6
    ]

    confused: list[tuple[str, str, int]] = []
    for true_label in q_labels:
        row = cm_q.loc[true_label]
        if row.sum() == 0:
            continue
        pred_label = row.idxmax()
        if pred_label != true_label:
            confused.append((true_label, pred_label, int(row[pred_label])))

    confused.sort(key=lambda x: -x[2])

    return {
        "quadrant_accuracy": float(accuracy_score(y_q_true, y_q_pred)),
        "quadrant_macro_f1": float(
            f1_score(y_q_true, y_q_pred, labels=q_labels, average="macro", zero_division=0)
        ),
        "block_accuracy": float(accuracy_score(y_b_true, y_b_pred)),
        "block_macro_f1": float(
            f1_score(y_b_true, y_b_pred, labels=b_labels, average="macro", zero_division=0)
        ),
        "joint_accuracy": float(sum(joint) / len(joint)) if joint else 0.0,
        "low_confidence_ratio": float(sum(low_conf) / len(low_conf)) if low_conf else 0.0,
        "confusion_matrix_quadrant": cm_q,
        "confusion_matrix_block": cm_b,
        "most_confused_pairs": confused[:15],
        "samples_evaluated": len(y_q_true),
    }


def optimize_ensemble_weights(
    clf: "TechRadarClassifier",
    df_val: pd.DataFrame,
    weight_grid: list[float] | None = None,
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    if weight_grid is None:
        weight_grid = [round(x * 0.1, 1) for x in range(1, 10)]

    best_q = 0.4
    best_b = 0.6
    best_q_score = -1.0
    best_b_score = -1.0

    for weight_rule in weight_grid:
        clf.set_ensemble_weights(quadrant=weight_rule)
        metrics_q = evaluate(df_val, clf)
        if metrics_q["quadrant_macro_f1"] > best_q_score:
            best_q_score = metrics_q["quadrant_macro_f1"]
            best_q = weight_rule

    for weight_rule in weight_grid:
        clf.set_ensemble_weights(block=weight_rule)
        metrics_b = evaluate(df_val, clf)
        if metrics_b["block_macro_f1"] > best_b_score:
            best_b_score = metrics_b["block_macro_f1"]
            best_b = weight_rule

    clf.set_ensemble_weights(quadrant=best_q, block=best_b)
    final_metrics = evaluate(df_val, clf)
    payload = {
        "optimized_weight_rule": round((best_q + best_b) / 2, 2),
        "optimized_weight_rule_quadrant": best_q,
        "optimized_weight_rule_block": best_b,
        "optimized_quadrant_macro_f1": round(best_q_score, 4),
        "optimized_block_macro_f1": round(best_b_score, 4),
        "optimized_macro_f1_avg": round((best_q_score + best_b_score) / 2, 4),
        "validation_joint_accuracy": round(final_metrics["joint_accuracy"], 4),
        "weight_grid": weight_grid,
    }
    target = Path(output_path or ENSEMBLE_WEIGHTS_PATH)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)
    return payload


def export_batch_to_excel(
    df_input: pd.DataFrame,
    output_path: str,
    clf: "TechRadarClassifier | None" = None,
) -> Path:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    if clf is None:
        from classifier import TechRadarClassifier

        clf = TechRadarClassifier(rebuild_prototypes=False)

    batch_input = df_input.copy()
    if "block" not in batch_input.columns:
        batch_input["block"] = None
    preds_rows = []
    for _, row in batch_input.iterrows():
        result = clf.classify(
            str(row.get("name", "") or ""),
            str(row.get("description", "") or ""),
            raw_block=row.get("block"),
            ring=row.get("ring"),
        )
        preds_rows.append(
            {
                "name": result.name,
                "quadrant": result.quadrant,
                "block": result.block,
                "quadrant_confidence": result.quadrant_confidence,
                "block_confidence": result.block_confidence,
                "quadrant_top3": result.quadrant_top3,
                "block_top3": result.block_top3,
                "classification_method": result.classification_method,
                "warnings": result.warnings,
            }
        )
    preds = pd.DataFrame(preds_rows)
    out = df_input.copy().reset_index(drop=True)
    for col in [
        "quadrant_pred",
        "block_pred",
        "quadrant_conf",
        "block_conf",
        "method",
        "warnings",
    ]:
        out[col] = None

    out["quadrant_pred"] = preds["quadrant"]
    out["block_pred"] = preds["block"]
    out["quadrant_conf"] = preds["quadrant_confidence"]
    out["block_conf"] = preds["block_confidence"]
    out["method"] = preds["classification_method"]
    out["warnings"] = preds["warnings"].apply(lambda x: "; ".join(x) if isinstance(x, list) else str(x))

    markup_cols = [
        "name",
        "description",
        "quadrant_pred",
        "block_pred",
        "quadrant_conf",
        "block_conf",
        "method",
        "warnings",
    ]
    if "quadrant_true" in out.columns or "quadrant" in out.columns:
        true_q = out.get("quadrant_true", out.get("quadrant"))
        if true_q is not None:
            out["quadrant_true"] = true_q
            markup_cols.insert(2, "quadrant_true")
    if "block_true" in out.columns or "block" in out.columns:
        true_b = out.get("block_true", out.get("block"))
        if true_b is not None:
            out["block_true"] = true_b
            markup_cols.insert(markup_cols.index("quadrant_pred") + 1, "block_true")

    top3_rows = []
    for idx, pred in preds.iterrows():
        row = {"name": pred["name"]}
        for i, (label, score) in enumerate(pred["quadrant_top3"][:3], start=1):
            row[f"quadrant_{i}"] = label
            row[f"quadrant_score_{i}"] = score
        for i in range(len(pred["quadrant_top3"]) + 1, 4):
            row[f"quadrant_{i}"] = ""
            row[f"quadrant_score_{i}"] = ""
        for i, (label, score) in enumerate(pred["block_top3"][:3], start=1):
            row[f"block_{i}"] = label
            row[f"block_score_{i}"] = score
        for i in range(len(pred["block_top3"]) + 1, 4):
            row[f"block_{i}"] = ""
            row[f"block_score_{i}"] = ""
        top3_rows.append(row)
    top3_df = pd.DataFrame(top3_rows)

    review_df = out[
        (out["quadrant_conf"] < 0.6)
        | (out["block_conf"] < 0.6)
        | out["warnings"].str.contains("НН-Спутник", na=False)
    ].copy()
    review_df["confidence_sum"] = review_df["quadrant_conf"] + review_df["block_conf"]
    review_df = review_df.sort_values("confidence_sum").drop(columns=["confidence_sum"])

    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)

    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)

    def confidence_fill(value: object) -> PatternFill:
        try:
            val = float(value)
        except (TypeError, ValueError):
            return red
        if val >= 0.8:
            return green
        if val >= 0.6:
            return yellow
        return red

    with pd.ExcelWriter(target, engine="openpyxl") as writer:
        out[markup_cols].to_excel(writer, sheet_name="Разметка", index=False)
        top3_df.to_excel(writer, sheet_name="Топ-3 кандидаты", index=False)
        review_df[markup_cols].to_excel(writer, sheet_name="Требуют проверки", index=False)

        has_truth = "quadrant_true" in out.columns or (
            "quadrant" in out.columns and "quadrant_pred" in out.columns
        )
        if has_truth:
            eval_input = out[["name", "description"]].copy()
            if "quadrant_true" in out.columns:
                eval_input["quadrant"] = out["quadrant_true"].map(canonical_quadrant)
            else:
                eval_input["quadrant"] = out["quadrant"].map(canonical_quadrant)
            if "block_true" in out.columns:
                eval_input["block"] = out["block_true"].map(canonical_block)
            else:
                eval_input["block"] = out["block"].map(canonical_block)
            eval_input = eval_input.dropna(subset=["quadrant", "block"])
            metrics = evaluate(eval_input, clf)
            metric_rows = [
                ("quadrant_accuracy", metrics["quadrant_accuracy"]),
                ("quadrant_macro_f1", metrics["quadrant_macro_f1"]),
                ("block_accuracy", metrics["block_accuracy"]),
                ("block_macro_f1", metrics["block_macro_f1"]),
                ("joint_accuracy", metrics["joint_accuracy"]),
            ]
            pd.DataFrame(metric_rows, columns=["metric", "value"]).to_excel(
                writer, sheet_name="Метрики", index=False, startrow=0
            )
            cm_q = metrics["confusion_matrix_quadrant"]
            cm_q.to_excel(writer, sheet_name="Метрики", startrow=len(metric_rows) + 2)
            start_row = len(metric_rows) + len(cm_q) + 4
            cm_b = metrics["confusion_matrix_block"]
            cm_b.to_excel(writer, sheet_name="Метрики", startrow=start_row)

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = header_font
            for col_idx in range(1, ws.max_column + 1):
                letter = get_column_letter(col_idx)
                max_len = 0
                for cell in ws[letter]:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[letter].width = min(max_len + 2, 60)

            if sheet_name == "Разметка":
                headers = {cell.value: cell.column for cell in ws[1]}
                for conf_name in ("quadrant_conf", "block_conf"):
                    if conf_name in headers:
                        col = headers[conf_name]
                        for row in range(2, ws.max_row + 1):
                            ws.cell(row=row, column=col).fill = confidence_fill(
                                ws.cell(row=row, column=col).value
                            )

    return target


def save_metrics_json(metrics: dict[str, Any], path: str | Path | None = None) -> Path:
    target = Path(path or ITERATION3_METRICS_PATH)
    target.parent.mkdir(parents=True, exist_ok=True)
    serializable: dict[str, Any] = {}
    for key, value in metrics.items():
        if isinstance(value, pd.DataFrame):
            serializable[key] = value.to_dict()
        elif isinstance(value, (np.floating, np.integer)):
            serializable[key] = float(value)
        else:
            serializable[key] = value
    with target.open("w", encoding="utf-8") as fh:
        json.dump(serializable, fh, ensure_ascii=False, indent=2, default=str)
    return target
